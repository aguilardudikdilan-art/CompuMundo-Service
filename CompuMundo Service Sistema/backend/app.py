
import os
import threading
from datetime import datetime
from flask import Flask, request, jsonify
from flask_cors import CORS
from openpyxl import Workbook, load_workbook
from filelock import FileLock

APP_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_PATH = os.path.join(APP_DIR, 'tickets.xlsx')
LOCK_PATH = DATA_PATH + '.lock'

COLUMNS = ['id','fechaISO','nombre','telefono','email','dispositivo','descripcion','prioridad','estado']

app = Flask(__name__)
CORS(app)

# Inicializa archivo Excel si no existe
def ensure_excel():
    if not os.path.exists(DATA_PATH):
        wb = Workbook()
        ws = wb.active
        ws.title = 'tickets'
        ws.append(COLUMNS)
        wb.save(DATA_PATH)

def read_all():
    ensure_excel()
    with FileLock(LOCK_PATH, timeout=10):
        wb = load_workbook(DATA_PATH)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
    if not rows:
        return []
    headers = rows[0]
    data = []
    for r in rows[1:]:
        if r and any(r):
            data.append({headers[i]: r[i] for i in range(len(headers))})
    return data

def append_row(record):
    ensure_excel()
    with FileLock(LOCK_PATH, timeout=10):
        wb = load_workbook(DATA_PATH)
        ws = wb.active
        ws.append([record.get(col, '') for col in COLUMNS])
        wb.save(DATA_PATH)
        wb.close()

def update_row(id_, parcial):
    ensure_excel()
    updated = False
    with FileLock(LOCK_PATH, timeout=10):
        wb = load_workbook(DATA_PATH)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        idx_id = headers.index('id')
        for row in ws.iter_rows(min_row=2):
            if str(row[idx_id].value) == str(id_):
                for k,v in parcial.items():
                    if k in COLUMNS:
                        cidx = headers.index(k)
                        row[cidx].value = v
                updated = True
                break
        if updated:
            wb.save(DATA_PATH)
        wb.close()
    return updated

@app.route('/')
def root():
    return jsonify({'ok': True, 'service': 'IriondoTech Tickets API', 'endpoints': ['/tickets']})

@app.route('/tickets', methods=['GET'])
def list_tickets():
    data = read_all()
    return jsonify({'ok': True, 'data': data})

@app.route('/tickets', methods=['POST'])
def create_ticket():
    body = request.get_json(force=True) or {}
    new_id = 'T' + datetime.utcnow().strftime('%Y%m%d%H%M%S%f')
    record = {
        'id': new_id,
        'fechaISO': datetime.utcnow().isoformat(),
        'nombre': body.get('nombre',''),
        'telefono': body.get('telefono',''),
        'email': body.get('email',''),
        'dispositivo': body.get('dispositivo',''),
        'descripcion': body.get('descripcion',''),
        'prioridad': body.get('prioridad','Media'),
        'estado': 'Nuevo'
    }
    append_row(record)
    return jsonify({'ok': True, 'id': new_id})

@app.route('/tickets/<id_>', methods=['PUT'])
def update_ticket(id_):
    body = request.get_json(force=True) or {}
    parcial = {}
    if 'estado' in body: parcial['estado'] = body['estado']
    if 'prioridad' in body: parcial['prioridad'] = body['prioridad']
    ok = update_row(id_, parcial)
    if not ok:
        return jsonify({'ok': False, 'error': 'ID no encontrado'}), 404
    return jsonify({'ok': True})

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 8000))
    ensure_excel()
    app.run(host='0.0.0.0', port=port, debug=True)
